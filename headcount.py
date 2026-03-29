"""
modules/headcount.py
Reconciles a monthly headcount file against the Periodics sheet.
"""

import re
import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from config import CONFIG

def _normalise(name):
    if pd.isna(name): return ""
    return re.sub(r"\s+", " ", str(name).strip().lower())

def _thin_border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def _style_header(cell, bg="1F4E79"):
    cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.fill      = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center",
                               wrap_text=True)
    cell.border    = _thin_border()

def _style_cell(cell, bg=None):
    cell.font      = Font(name="Arial", size=10)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    cell.border    = _thin_border()
    if bg:
        cell.fill = PatternFill("solid", start_color=bg)


class HeadcountReconciler:

    def __init__(self, ui=None):
        self.ui    = ui
        self.path  = CONFIG["EXCEL_PATH"]
        self.sheet = CONFIG["PERIODICS_SHEET"]
        self.new_sheet = CONFIG["NEW_EMP_SHEET"]

    def reconcile(self, hc_path):
        if self.ui:
            self.ui.info(f"Reconciling: {os.path.basename(hc_path)}")

        try:
            hc_df = (pd.read_csv(hc_path)
                     if hc_path.endswith(".csv")
                     else pd.read_excel(hc_path))
        except Exception as e:
            if self.ui: self.ui.error(f"Cannot read headcount: {e}")
            return 0, 0

        if "Personnel Names" not in hc_df.columns:
            if self.ui:
                self.ui.warn("'Personnel Names' column missing — skipping")
            return 0, 0

        hc_df["_norm"] = hc_df["Personnel Names"].apply(_normalise)

        try:
            per_df = pd.read_excel(self.path, sheet_name=self.sheet)
        except Exception as e:
            if self.ui: self.ui.error(f"Cannot read Periodics sheet: {e}")
            return 0, 0

        if "Personnel Names" not in per_df.columns:
            if self.ui:
                self.ui.warn("'Personnel Names' missing in Periodics — skipping")
            return 0, 0

        per_df["_norm"] = per_df["Personnel Names"].apply(_normalise)

        hc_names     = set(hc_df["_norm"]) - {""}
        per_names    = set(per_df["_norm"]) - {""}
        new_names    = hc_names  - per_names
        exited_names = per_names - hc_names

        if self.ui:
            self.ui.result("New employees",    len(new_names))
            self.ui.result("Possible leavers", len(exited_names))

        today_str = datetime.today().strftime("%d-%b-%Y")

        # Flag leavers
        if exited_names:
            for col in ["UpdateStatus", "UpdateDate"]:
                if col not in per_df.columns:
                    per_df[col] = None
            mask = per_df["_norm"].isin(exited_names)
            per_df.loc[mask, "UpdateStatus"] = "Exited"
            per_df.loc[mask, "UpdateDate"]   = today_str
            clean = per_df.drop(columns=["_norm"])
            with pd.ExcelWriter(self.path, engine="openpyxl", mode="a",
                                if_sheet_exists="replace") as writer:
                clean.to_excel(writer, sheet_name=self.sheet, index=False)
            if self.ui:
                self.ui.warn(f"Flagged {len(exited_names)} employee(s) as Exited")

        # Write new employees sheet
        new_rows = hc_df[hc_df["_norm"].isin(new_names)].copy()
        new_rows = new_rows.drop(columns=["_norm"])
        new_rows["Flagged On"]   = today_str
        new_rows["Action Taken"] = "Pending — Medical Exam Not Yet Booked"
        self._write_new_emp_sheet(new_rows)

        if self.ui:
            self.ui.success(
                f"{len(new_names)} new employee(s) written to "
                f"'{self.new_sheet}' sheet"
            )

        return len(new_names), len(exited_names)

    def _write_new_emp_sheet(self, df):
        wb   = load_workbook(self.path)
        if self.new_sheet in wb.sheetnames:
            del wb[self.new_sheet]
        ws   = wb.create_sheet(self.new_sheet)
        cols = list(df.columns)

        for c_idx, col in enumerate(cols, 1):
            _style_header(ws.cell(row=1, column=c_idx, value=col))
        ws.row_dimensions[1].height = 28

        for r_idx, (_, row) in enumerate(df.iterrows(), 2):
            for c_idx, col in enumerate(cols, 1):
                cell = ws.cell(row=r_idx, column=c_idx, value=row[col])
                _style_cell(cell, bg="FFF2CC")

        for c_idx, col in enumerate(cols, 1):
            vals = [str(col)] + [str(df.iloc[r][col])
                                 for r in range(len(df))]
            ws.column_dimensions[get_column_letter(c_idx)].width = \
                min(max(len(v) for v in vals) + 4, 40)

        ws.freeze_panes = "A2"
        wb.save(self.path)
